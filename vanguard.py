import pandas as pd
from servant import clean, get_month, get_pension_type, from_df_to_dict
import openpyxl
import openpyxl.utils


class Vanguard:
    def __init__(self, file_mzdy, file_pracov):
        self.mzdy = file_mzdy
        self.pracov = file_pracov
        self.dataframe = self.prep_df

    @property
    def prep_df(self):
        """ Preparation of data to dictionary. Uses pandas library to open, sort and merge data from csvs."""
        mzdy = pd.read_csv(self.mzdy, encoding='cp1250').applymap(clean)
        mzdy['Fare'] = mzdy[['Davky1', 'Davky2']].sum(axis=1, skipna=True)
        mzdy['Payout'] = mzdy[['Zamest', 'HrubaMzda', 'iNemoc']].sum(axis=1, skipna=True)
        mzdy['RokMes'] = mzdy['RokMes'].map(get_month)

        pracov = pd.read_csv(self.pracov, encoding='cp1250').applymap(clean)
        pracov['PensionType'] = pracov['TypDuch'].map(get_pension_type)

        data = pd.merge(mzdy, pracov, on='RodCislo', suffixes=('', '_y'))
        data = data.drop(['Kat_y', 'Kod_y', 'Jmeno30', 'Davky1', 'Davky2', 'Zamest', 'HrubaMzda', 'iNemoc'], axis=1)

        return data

    @property
    def loader(self):
        """ { idnum1: { name: 'abc', date: { month: { fare: 1234, payout: 9886 } } other_data: ... }, idnum2: ..., } """
        # Dictionary with employees data from exported CSVs merged into single object.
        merged = from_df_to_dict(self.dataframe, True)
        x = XScout('tables/jmenny_seznam_2022_09_27 Fiala.xlsx')

        employee_list = x.employee_list()

        return employee_list, merged, x.range


class XScout:
    def __init__(self, spreadsheet):
        self.wb = openpyxl.load_workbook(spreadsheet)
        self.range = self.spread

    def employee_list(self):
        """ Returns list of people present on spreadsheet. Each sheet has its own dictionary with person:row kw pair"""
        people = []
        for sheet_index, ws in enumerate(self.wb.worksheets[1:3]):
            sheet_ids = {}
            for row in range(self.range[sheet_index][0], self.range[sheet_index][1] + 1):
                data_row = clean(ws.cell(row, 4).value)
                sheet_ids[data_row] = row
            people.append(sheet_ids)

        return people

    def first_row(self, sheet_number):
        ws = self.wb.worksheets[sheet_number]
        row = 1
        while True:
            if ws.cell(row, 1).value == 1:
                return row
            row += 1

    def last_row(self, sheet_number):
        ws = self.wb.worksheets[sheet_number]
        row = self.first_row(sheet_number)

        while True:
            if not ws.cell(row, 2).value:
                return row - 1
            row += 1

    @property
    def spread(self):
        s = []
        for sheet_number in range(1, 3):
            spread = (self.first_row(sheet_number), self.last_row(sheet_number))
            s.append(spread)
        return s



