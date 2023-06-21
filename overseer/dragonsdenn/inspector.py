import re
import numpy as np
from . import servant
import openpyxl
from openpyxl.utils import column_index_from_string

from pathlib import Path
import pandas as pd
import platform

pd.set_option('display.max_rows', 100)


def make_df(filename, sheetname, rows_to_skip, use_cols, nrows):
    df = pd.read_excel(
        filename,
        sheet_name=sheetname,
        header=None,
        skiprows=rows_to_skip,
        usecols=use_cols,
        nrows=nrows - rows_to_skip - 1,
        index_col=None
    )
    # Update the names of columns and add unique number to columns with the same name
    column_names = []
    count_dict = {}
    for column_name in df.iloc[0].str.replace(r'\s+', ' ', regex=True):
        if column_name in count_dict:
            count_dict[column_name] += 1
            column_names.append(column_name + '_' + str(count_dict[column_name]))
        else:
            count_dict[column_name] = 1
            column_names.append(column_name)
    df.columns = column_names
    #
    df = df.drop([0])
    return df


class Inspector:
    def __init__(self, up_spreadsheet, lo_spreadsheet):
        self.file_up = up_spreadsheet
        self.file_lo = lo_spreadsheet
        self.wb_up, self.wb_lo = self.is_readable()
        self.q = self.quarter
        self.df_up = self.load_df_up
        self.df_lo = self.load_df_loc

    def is_readable(self):
        ''' 
        TODO
        check if libreoffice or ms office is installed.
        ''' 

        wb_up = openpyxl.load_workbook(self.file_up, data_only=True)
        ws = wb_up.worksheets[0]

        if ws.cell(21, 7).value is None:
            if platform.system() == 'Windows':
                servant.saveas_excel(self.file_up)
            else:
                servant.saveas_libreoffice(self.file_up)
            
            wb_up = openpyxl.load_workbook(self.file_up, data_only=True)

        wb_lo = servant.unlock(self.file_lo, '13881744', data_only=True)
        ws = wb_lo.worksheets[-2]

        if ws.cell(3, 2).value is None:
            if platform.system() == 'Windows':
                servant.saveas_excel(self.file_lo)
            else:
                servant.saveas_libreoffice(self.file_lo)

            wb_lo = servant.unlock(self.file_lo, '13881744', data_only=True)

        return wb_up, wb_lo

    @property
    def refund_up(self):

        ws = self.wb_up.worksheets[0]
        ref_val = ws.cell(21, 7).value

        return ref_val

    @property
    def refund_lo(self):
        ws = self.wb_lo.worksheets[-3]

        if self.q:
            ref_val = ws.cell(self.q + 2, column_index_from_string('R')).value
            return ref_val
        else:
            print(f'Quarter is not defined. Check file {self.file_up}, cell D4')

    @property
    def quarter(self):
        ws = self.wb_up.worksheets[0]
        quarter = ws.cell(6, 4).value

        try:
            return int(quarter)
        except ValueError:
            print(f'Quarter is not integer type. Fix the value in {self.file_up} in cell D6')
            return False

    def refunds_are_equal(self):
        rup = self.refund_up
        rlo = self.refund_lo
        if rup == rlo:
            return True
        else:
            print('Refunds are not equal')
            print(f'RUP: {rup}, RLO: {rlo}, difference: {rup - rlo}')
            return False

    def refund_for_month_up(self, month=0):
        total = self.df_up.iloc[:, 11 + (month * 5):14 + (month * 5)].apply(pd.to_numeric, errors='coerce').sum()
        return total.sum()

    def refund_for_month_lo(self, month=0):
        month_index = self.q * 3 - 2 + month
        total = self.df_lo[f'refundace{month_index}'].apply(pd.to_numeric, errors='coerce').sum()
        return total

    @property
    def faulty_months(self):
        if not self.refunds_are_equal():
            faulty_months = []
            refunds_up = [self.refund_for_month_up(refund) for refund in range(3)]
            refunds_lo = [self.refund_for_month_lo(refund) for refund in range(3)]

            for month_index, month in enumerate(zip(refunds_up, refunds_lo)):
                if month[0] != month[1]:
                    faulty_months.append(month_index)

            for month in faulty_months:
                print(refunds_up[month], refunds_lo[month])

            return faulty_months
        return False

    @property
    def load_df_loc(self):
        wb = openpyxl.load_workbook(self.file_lo, read_only=False, data_only=True)
        sheetnames = wb.worksheets[:-4]
        table_indexes = [i for i in range(1, 9)]

        sheet_dfs = []

        for table_index, ws in zip(table_indexes, sheetnames):
            tab = ws.tables[f'Tabulka{table_index}']
            table_range = tab.ref
            rows = int(re.search(r'\d*$', table_range).group())

            columns = re.sub(r'\d', '', table_range)

            df = make_df(self.file_lo, ws.title, 1, columns, rows)
            sheet_dfs.append(df)

        big_df = pd.concat([*sheet_dfs], ignore_index=True).sort_values('Jméno')
        big_df.dropna(subset=['Jméno'], inplace=True)
        big_df['idx'] = range(1, len(big_df) + 1)
        big_df.set_index('idx', inplace=True)
        return big_df

    @property
    def load_df_up(self):
        ws = self.wb_up.worksheets[1]
        sheet_name = ws.title
        row = 13

        # Get last line
        while True:
            if not ws.cell(row, 2).value:
                break
            row += 1

        df = make_df(self.file_up, sheet_name, 11, 'A:X', row)

        df['refundace0'] = df.iloc[:, 11:14].apply(pd.to_numeric, errors='coerce').sum(axis=1)
        df['refundace1'] = df.iloc[:, 16:19].apply(pd.to_numeric, errors='coerce').sum(axis=1)
        df['refundace2'] = df.iloc[:, 21:24].apply(pd.to_numeric, errors='coerce').sum(axis=1)
        df['jmeno'] = df['Příjmení'] + ' ' + df['Jméno']
        df['idx'] = range(1, len(df) + 1)
        df.set_index('idx', inplace=True)
        return df

    @property
    def combined(self):
        merged_df = pd.merge(self.df_lo[['Jméno', 'refundace7', 'refundace8', 'refundace9']],
                             self.df_up[['jmeno', 'refundace0', 'refundace1', 'refundace2']]
                             , left_on='Jméno', right_on='jmeno', how='outer').replace(np.NAN, 0)
        return merged_df

    def check_faulty_months(self):
        combined = self.combined
        for month in self.faulty_months:
            print(month)
            is_not_equal = combined.iloc[:, 1 + month] != combined.iloc[:, 5 + month]
            values = combined.loc[is_not_equal, combined.columns[[0, 1 + month, 5 + month]]]
            print(values)


def main():
    file_up = Path.home() / '.amn/temp-up.xlsx'
    file_lo = Path.home() / '.amn/temp.xlsx'

    if all((Path(file_up).exists(), Path(file_lo).exists())):
        inspector = Inspector(file_up, file_lo)
        # print(inspector.faulty_months)
        # print(inspector.df_up[['Příjmení', 'refundace0', 'refundace1', 'refundace2']])
        # print()
        # print(inspector.df_up.columns)
        # print(inspector.df_lo['Jméno'])
        # print(inspector.df_up['jmeno'])
        # print(inspector.combined)
        inspector.check_faulty_months()
        # print(inspector.df_lo.columns)
        # print(inspector.df_lo['Jméno'].isin(inspector.df_up['jmeno']), inspector.df_lo['Jméno'], inspector.df_up['jmeno'])
    else:
        print('File does not exist')
