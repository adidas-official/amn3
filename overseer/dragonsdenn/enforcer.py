# Importing libraries
from pathlib import Path
from itertools import chain
from datetime import datetime
from openpyxl.utils import get_column_letter
import traceback, sys

# Local imports
from . import servant
from .vanguard import Assembler
from .courier import logger, message as msg, msgrow


def write_to_list(idnum, items, row_num, worksheet, new=False) -> dict:
    """
    Args:
        idnum (str): The ID number of the servant.
        items (dict): A dictionary of the servant's data.
        row_num (int): The row number to add the data.
        worksheet (Worksheet): The worksheet to add the data.
        new (bool): Whether or not the servant is new.

    Returns:
        Dict: A dictionary of the data to be sent in the message.
    """
    data4message = {}
    
    if worksheet.title.startswith('2'):
        num = 0
        offset = 1
    else:
        num = 1
        offset = 0

    for month, payout_data in items["Date"].items():
        m = (int(month.split('.')[0]) + 2) % 3
        shift = 14 - m
        if items["PensionType"]:
            month_col_status = servant.column_map[num][m]
            worksheet.cell(row_num, month_col_status).value = items["PensionType"]

        month_col_payout = servant.column_map[num][m] + offset
        worksheet.cell(row_num, month_col_payout).value = payout_data["Payout"]
        if payout_data["Payout"] > 0:
            worksheet.cell(row_num, month_col_payout + shift).value = f'=IF({get_column_letter(month_col_payout)}{row_num}<>"";14200-{get_column_letter(month_col_payout + 1)}{row_num};0)'

        if num == 0:
            worksheet.cell(row_num, 6).value = items["EndEmployment"]

        if new:
            lastname, firstname = servant.split_name(items["Name"])
            worksheet.cell(row_num, 2).value = lastname
            worksheet.cell(row_num, 3).value = firstname
            if num == 0:
                worksheet.cell(row_num, 4).value = idnum
                worksheet.cell(row_num, 5).value = items["StartEmployment"]
                worksheet.cell(row_num, 7).value = items["InsCode"]
                worksheet.cell(row_num, 8).value = items["PensionStart"]
            elif num == 1:
                worksheet.cell(row_num, 4).value = idnum[:6] + '/' + idnum[6:]
                worksheet.cell(row_num, 5).value = '-\'\'-'
                worksheet.cell(row_num, 6).value = 'PA'
                worksheet.cell(row_num, 7).value = '100%'

        data4message.setdefault(str(month), {'payout': payout_data["Payout"], "cell": get_column_letter(month_col_payout) + str(row_num)})

    logger.info(msg(items["Name"], 1, data4message, new))
    return msgrow(items["Name"], 1, data4message, new)


class Enforcer:
    def __init__(self, dataset):
        self.data_up = dataset[0][0]
        self.data_lo = dataset[0][1]
        self.merged_up = dataset[1][0]
        self.merged_lo = dataset[1][1]
        self.scout = dataset[-2]
        self.last_row = [row[1] + 1 for row in self.scout.range]
        self.quarter = dataset[-1]
    
    def make_home_dir(self):
        home_dir = Path.home()
        amn_dir = home_dir / '.amn'
        if not amn_dir.exists():
            amn_dir.mkdir()
        return amn_dir

    def write_data_lo(self) -> None:
        """Write employees' data to the LO file"""
        logger.info("-" * 64)
        logger.info("Writing data to LO file")
        outdir = Path(self.make_home_dir())
        wb = self.scout.wb_lo
        self.write_new_emps_lo(wb)
        for person, data in self.merged_lo.items():
            for i, sheet_data in enumerate(self.data_lo):
                if i < 3:
                    fare_shift = 4
                else:
                    fare_shift = 2

                if person in sheet_data:
                    data4msg = {}
                    message = f'Sheet: {i}, Line: {sheet_data[person]}, Person: {person}, {data["Code"]}, {data["Cat"]}'
                    ws = wb.worksheets[i]
                    for date, money in data['Date'].items():
                        col = self.scout.get_month(date, i)
                        fare_col = col + fare_shift
                        col_letter = get_column_letter(col)
                        fare_letter = get_column_letter(fare_col)
                        message += f'\n- {col_letter}{sheet_data[person]}: {money["Payout"]}'
                        ws.cell(sheet_data[person], col).value = money["Payout"]
                        if money["Fare"]:
                            message += f', {fare_letter}{sheet_data[person]}: {money["Fare"]}'
                            ws.cell(sheet_data[person], fare_col).value = money["Fare"]
                        data4msg.setdefault(str(date), {'payout': money["Payout"], "cell": get_column_letter(col) + str(sheet_data[person])})
                        logger.info(msg(person, i, data4msg))
                    # logger.info(message)
        wb.save(outdir / 'temp.xlsx')

    def get_new_emps(self) -> list:
        data_for_month = set(self.merged_lo.keys())
        names_in_xlsx = set(chain.from_iterable(d.keys() for d in self.data_lo))
        return [self.merged_lo[name[:20]] for name in data_for_month.difference(names_in_xlsx)]

    def write_new_emps_lo(self, wb) -> None:
        """ Write new employees to the LO file"""
        last_lines = self.get_last_rows()

        for emp in self.get_new_emps():
            sheet_index = servant.get_sheet_by_emp_data(emp['Code'], emp['Cat'])

            ws = wb.worksheets[sheet_index]
            servant.insert_row(ws, last_lines[sheet_index] + 1)
            ws.cell(last_lines[sheet_index] + 1, 2).value = emp["Name"]
            message = f'New person: {emp["Name"]}, sheet: {sheet_index}, line: {last_lines[sheet_index] + 1}'
            fare = 0

            if emp["PensionType"]:
                pension = "With pension"
                ws.cell(last_lines[sheet_index] + 1, 1).value = 1
            else:
                pension = "without pension"
                ws.cell(last_lines[sheet_index] + 1, 1).value = 0

            if sheet_index < 3:
                fare_shift = 4
            else:
                fare_shift = 2

            for date, money in emp['Date'].items():
                col = self.scout.get_month(date, sheet_index)
                fare_col = col + fare_shift

                ws.cell(last_lines[sheet_index] + 1, col).value = money["Payout"]

                if money["Fare"]:
                    fare = money["Fare"]
                    ws.cell(last_lines[sheet_index] + 1, fare_col).value = money["Fare"]
                message += f'- Month: {date}, Payout: {money["Payout"]}, Fare: {fare}, {pension}'
                # logger.info(message)

            last_lines[sheet_index] += 1

    def write_data(self) -> list:
        """Write employees' data to the UP file, including filling new employees"""
        return_msg = []

        logger.info(msg)

        outdir = Path(self.make_home_dir())
        wb = self.scout.wb_up
        ws = wb.worksheets[0]
        ws.cell(6, 4).value = self.quarter
        ws.cell(6, 9).value = datetime.now().year
        ws.cell(30, 5).value = datetime.now().strftime('%d.%m.%Y')

        for person, data in self.merged_up.items():
            if person in self.data_up[0]:
                ws = wb.worksheets[1]
                return_msg.append(write_to_list(person, data, self.data_up[0][person], ws))

            elif person in self.data_up[1]:
                ws = wb.worksheets[2]
                return_msg.append(write_to_list(person, data, self.data_up[1][person], ws))
            else:
                if data['PensionType'] != '':
                    ws = wb.worksheets[1]
                    return_msg.append(write_to_list(person, data, self.last_row[0], ws, new=True))
                    self.last_row[0] += 1
                else:
                    ws = wb.worksheets[2]
                    return_msg.append(write_to_list(person, data, self.last_row[1], ws, new=True))
                    self.last_row[1] += 1
        wb.save(outdir / 'temp-up.xlsx')
        return return_msg

    def get_last_rows(self) -> list:
        last_rows = [self.scout.last_row_lo(i) for i in range(8)]
        return last_rows

def main(wages, employees):
    # if all ok, send success message
    # raise exceptions if something goes wrong
    # on exception, send error message

    try:
        logger.info('Starting')
        return_msg = []

        vanguard = Assembler(data_mzdy=wages, data_pracov=employees)
        enforcer = Enforcer(vanguard.loader)
        return_msg.append(enforcer.write_data())
        logger.debug(enforcer.write_data_lo())

        logger.info('Done')
        return return_msg
    except Exception as e:
        logger.error(traceback.print_exc(file=sys.stdout))
        return 'Exception occurred. Check logs'