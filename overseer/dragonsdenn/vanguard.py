# Importing libraries
import pandas as pd
import openpyxl
import openpyxl.utils
from pathlib import Path
from io import StringIO

# Local imports
from dragonsdenn import servant, paths
from dragonsdenn.courier import logger



class Assembler:
    def __init__(self, data_mzdy, data_pracov, loc_table, up_table):
        self.mzdy = StringIO(data_mzdy)
        self.pracov = StringIO(data_pracov)
        self.dataframe = self.prep_df
        self.loc_table = loc_table
        self.up_table = up_table

    @property
    def prep_df(self):
        """ Preparation of data to dictionary. Uses pandas library to open, sort and merge data from csvs."""
        logger.info('Preparing data.')

        # Load data
        try:
            mzdy = pd.read_csv(self.mzdy, encoding='cp1250').applymap(servant.clean)
        except FileNotFoundError:
            logger.error('File not found: {}'.format(self.mzdy))
            raise
        except UnicodeDecodeError:
            logger.error('File is not in Windows-1250 encoding: {}'.format(self.mzdy))
            raise
        pracov = pd.read_csv(self.pracov, encoding='cp1250').applymap(servant.clean)
        logger.info('Data from input csv files loaded into dataframe.')

        # Calculate additional columns
        if not mzdy.empty:
            self._mzdy = mzdy
            logger.info('Calcutating values for additional columns.')
            mzdy['Fare'] = mzdy[['Davky1', 'Davky2']].sum(axis=1, skipna=True)
            mzdy['Payout'] = mzdy[['Zamest', 'HrubaMzda', 'iNemoc']].sum(axis=1, skipna=True)
            mzdy['RokMes'] = mzdy['RokMes'].map(servant.get_month)
            pracov['PensionType'] = pracov['TypDuch'].map(servant.get_pension_type)
            logger.info('Additional columns calculated.')
        else:
            logger.info('Mzdy is empty.')

        # Join data
        data = pd.merge(mzdy, pracov, on='RodCislo', suffixes=('', '_y'))
        data = data.drop(['Kat_y', 'Kod_y', 'Jmeno30', 'Davky1', 'Davky2', 'Zamest', 'HrubaMzda', 'iNemoc'], axis=1)
        if data.shape[0] < 1:
            logger.error('No data.')
            return
        logger.info('Data joined.')

        return data

    @property
    def loader(self) -> tuple:
        """ Running all important pieces together. Creating list of employees in spreadsheets. """

        logger.info('Loading data to Enforcer.')
        # Converting dataframe to dictionary
        merged_lists = (servant.from_df_to_dict(self.dataframe, True, 'RodCislo'),
                        servant.from_df_to_dict(self.dataframe, False, 'JmenoS'))
        logger.info('Data converted to dictionary.')

        logger.info('Uploading spreadsheets to Enforcer.')
        loc = servant.save_uploaded_file(self.loc_table, Path(paths.TABLES_PATH))
        up = servant.save_uploaded_file(self.up_table, Path(paths.TABLES_PATH))

        # Creating list of all employees from scout
        scout = Scout(spreadsheet2=loc, spreadsheet1=up)
        cartographer = Cartographer(up, loc)

        employee_lists = (scout.employee_list_up(), scout.employee_list_lo())

        return employee_lists, merged_lists, scout, servant.get_q(self.dataframe), (loc, up), cartographer.lenght_of_months


class Scout:
    '''
    Finds people in spreadsheets. Creates dictionary with data about people in sheets.
    '''
    def __init__(self, spreadsheet1, spreadsheet2):
        self.wb_up = openpyxl.load_workbook(spreadsheet1)
        self.wb_lo = servant.unlock(spreadsheet2, '13881744')
        self.range = self.spread

    def employee_list_up(self) -> list:
        """Returns list of people present on spreadsheet. Each sheet has its own dictionary with person:row kw pair"""
        people = []
        # Loop through each worksheet in the workbook
        for sheet_index, ws in enumerate(self.wb_up.worksheets[1:3]):
            sheet_ids = {}
            for row in range(self.range[sheet_index][0], self.range[sheet_index][1] + 1):
                try:
                    # Get the data from the cell in ID column
                    data_row = servant.clean(ws.cell(row, 4).value)
                    sheet_ids[data_row] = row
                except AttributeError:
                    logger.error('AttributeError in employee_list_up')
                    pass
            people.append(sheet_ids)
        # Return the list of dicrionaries with people
        return people

    def first_row(self, sheet_number) -> int:
        """Returns number of first row with data in sheet."""
        ws = self.wb_up.worksheets[sheet_number]
        row = 1
        while True:
            if ws.cell(row, 1).value == 1:
                return row
            row += 1

    def last_row(self, sheet_number) -> int:
        """Returns number of last row with data in sheet."""
        ws = self.wb_up.worksheets[sheet_number]
        row = self.first_row(sheet_number)

        while True:
            if not ws.cell(row, 2).value:
                return row - 1
            row += 1

    def last_row_lo(self, sheet_number) -> int:
        """Returns number of last row with data in sheet."""
        if self.wb_lo:
            ws = self.wb_lo.worksheets[sheet_number]
        else:
            return False
        row = 3

        while True:
            if not ws.cell(row, 2).value:
                return row - 1
            row += 1

    def get_last_rows(self) -> list:
        if self.wb_lo:
            l = len(self.wb_lo.worksheets) -2
            logger.debug(f'Number of sheets: {l}')
            last_rows = [self.last_row_lo(i) for i in range(l)]
            return last_rows
        return []

    @property
    def spread(self) -> list:
        """Returns list of tuples with first and last row of each sheet."""
        s = []
        for sheet_number in range(1, 3):
            spread = (self.first_row(sheet_number), self.last_row(sheet_number))
            s.append(spread)
        return s

    def employee_list_lo(self):
        """ Returns list of people present on spreadsheet. Each sheet has its own dictionary with person:row kw pair"""
        if not self.wb_lo:
            logger.error('No workbook')
            return []
        people = []
        for ws in self.wb_lo.worksheets[:-2]:
            row = 3
            sheet_names = {}
            while True:
                name = str(ws.cell(row, 2).value)[:20]
                if name == '[ENDBLOCK]':
                    break

                if name:
                    sheet_names[name] = row
                row += 1
            
            people.append(sheet_names)

        return people

    def get_month(self, date, sheet_num):
        """ Gets month column index in local table from date in employee data object """

        if not self.wb_lo:
            logger.error('No workbook in get_month')
            return None

        # Counter for keeping track of curent column index
        counter = 3
        # Number of how many times merged cell was found
        m = 0
        month_num = int(date.split('.')[0])
        cell = False
        # If date is not in format 'M.YYYY' or 'M', return None
        try:
            month_num = int(date.split('.')[0])
        except IndexError:
            logger.error(f'IndexError in get_month. Date: {date}')
            return None
        except ValueError:
            logger.error(f'ValueError in get_month. Date: {date}')
            return None

        while True:
            cell = self.wb_lo.worksheets[sheet_num].cell(row=1, column=counter)

            if type(cell).__name__ == 'Cell' and cell.value:
                m += 1

            if m == month_num:
                return cell.column

            counter += 1


class Cartographer(Scout):
    # get xlsx files
    def __init__(self, spreadsheet1, spreadsheet2):
        super().__init__(spreadsheet1, spreadsheet2)
        self.first_row = 3
        self.first_column = 3

    # get first/last row

    # map months for each sheet
    @property
    def lenght_of_months(self) -> list:

        # each item in list is number of columns, that are one month
        cols_of_month = []

        for sheet in self.wb_lo.sheetnames[:-2]:
            ws = self.wb_lo[sheet]

            # initial column
            col_num = self.first_column
            while True:
                cell = ws.cell(row=1, column=col_num + 1)
                
                if type(cell).__name__ == 'Cell' and cell.value:
                    col_len = cell.column - self.first_column
                    cols_of_month.append(col_len)
                    # logger.debug(f'Month number #3 starts at column: {col_len * 2 + self.first_column}')
                    break
                    
                col_num += 1
                
        return cols_of_month

    def __str__(self):
        return "Hello, World!"


def run_test():
    logger.debug('Cartographer started')
    loc = Path(paths.TABLES_PATH) / 'temp.xlsx'
    up = Path(paths.TABLES_PATH) / 'temp-up.xlsx'

    ctg = Cartographer(up, loc)

    if not ctg:
        logger.debug('Cartographer failed.')
        return False

    logger.debug(ctg.lenght_of_months)

    return {'status':'Cartographer finished mapping.'}