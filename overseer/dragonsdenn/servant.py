import csv
import math
import re
import subprocess
import time
import zipfile
from io import BytesIO

import msoffcrypto
import openpyxl
import pyautogui
from openpyxl.utils import get_column_letter

from copy import copy
from shutil import which
from pathlib import Path
import platform
import tempfile
from . import paths
from .courier import logger
import os


column_map = [
    {
        0: 10,
        1: 15,
        2: 20
    },
    {
        0: 9,
        1: 10,
        2: 11
    }
]


def get_q(dataframe):
    dates = dataframe['RokMes'].unique()
    months = [int(month.split('.')[0]) for month in dates]
    nums = [math.ceil(i / 3) for i in months]
    result = all([x == nums[0] for x in nums])
    if result:
        return nums[0]


def unlock(filename, password=None, data_only=False):
    try:
        wb = openpyxl.load_workbook(filename, data_only=data_only, read_only=False)
        return wb
    except zipfile.BadZipFile:
        pass

    try:
        with open(filename, 'rb') as f:
            office_file = msoffcrypto.OfficeFile(f)
            if password:
                office_file.load_key(password=password)
            decrypted_wb = BytesIO()
            office_file.decrypt(decrypted_wb)
            wb = openpyxl.load_workbook(decrypted_wb, data_only=data_only, read_only=False)
            return wb
    except (msoffcrypto.exceptions.FileFormatError, ValueError):
        pass


def clean(text) -> str:
    if type(text) == str:
        return text.replace('\'', '').replace('/', '').strip()
    return text


def get_pension_type(pentype) -> str:
    if 'invalidní 1.' in pentype:
        return 'OZP12'
    elif 'invalidní 3.' in pentype:
        return 'TZP'
    elif 'zdravotně' in pentype:
        return 'OZZ'
    else:
        return ''


def split_name(name):
    fullname = name.split(' ')
    firstname = fullname[-1]
    lastname = ' '.join(fullname[:-1])
    return lastname, firstname


def get_sheet_by_emp_data(workplace, emp_status) -> int:
    if emp_status == 'DPP':
        return 8
    elif emp_status == 'U':
        return 9

    workplaces = {
        'Sklad': 0,
        'Admin': 0,
        'Jedna': 0,
        'Dílna': 1,
        'Prode': 2,
        'ÚP So': 3,
        'ÚP Ch': 4,
        'KÚ Ch': 5,
        'KÚ': 6,
        'Škola': 7,
        'KÚ ÚP': 6,
        '': 8
    }

    return workplaces[workplace]


def get_month(month) -> str:
    m = month.split('.')
    return '.'.join((m[1], m[2]))


def get_ins_code(code):
    with open(Path(paths.INSURANCE_PATH) / 'Fiala_insurance_codes.csv', 'r') as f:
        codes = csv.reader(f)
        data = dict(codes)
        data = {clean(k): clean(v) for k, v in data.items()}
        if code in data:
            return data[code]


def from_df_to_dict(df, filt=False, pk='RodCislo'):
    """Passes the dataframe to dictionary for later parsing."""
    people = {}
    if filt:
        filtered = (~df['Kat'].str.contains('U'))
        df = df[filtered]

    # Convert dataframe to dictionary in form like this:
    # {idnum: {'name': str, ... ,'date: {list of dates with payments}}, idnum2: {...}, }
    for _, row in df.iterrows():
        people.setdefault(row[pk], {
            'Name': row['JmenoS'],
            'Code': row['Kod'],
            'Cat': row['Kat'],
            'InsCode': get_ins_code(row['CisPoj']),
            'Date': {},
            'StartEmployment': row['VstupDoZam'],
            'EndEmployment': row['UkonceniZam'],
            'PensionType': row['PensionType'],
            'PensionStart': row['DuchOd']
        })
        people[row[pk]]['Date'].setdefault(row['RokMes'], {
            'Fare': int(row['Fare']), 'Payout': int(row['Payout'])
        })
    return people


def insert_row(worksheet, row_index):
    # define range of cells to move
    max_col = get_column_letter(worksheet.max_column)
    max_row = worksheet.max_row

    table_name, table_range = worksheet.tables.items()[0]
    last_row = int(re.search(r'\d*$', table_range).group())
    new_range = re.sub(r'\d*$', str(int(last_row) + 1), table_range, count=1)

    # Move all cells from the row bellow the insertion point, including styles and translated formulas
    worksheet.move_range(f"A{str(row_index)}:{max_col}{max_row}", rows=1, cols=0, translate=True)

    # Define new range of table
    worksheet.tables[table_name].ref = new_range

    pattern = re.compile(r'(?<=[A-Z])(\d+)')

    # Copy all cells from the row above the insertion point, including formatting and formulas
    for col in range(1, worksheet.max_column + 1):
        cell_above = worksheet.cell(row=row_index - 1, column=col)
        cell_to_copy = worksheet.cell(row=row_index, column=col)
        cell_value = cell_above.value
        cell_to_copy.value = cell_value

        if cell_value and type(cell_value) is not int:
            new_cell_value = re.sub(pattern, str(row_index), str(cell_value))
            cell_to_copy.value = new_cell_value

        cell_to_copy.number_format = cell_above.number_format
        cell_to_copy.font = copy(cell_above.font)
        if cell_above.has_style:
            cell_to_copy.border = copy(cell_above.border)
            cell_to_copy.fill = copy(cell_above.fill)
            cell_to_copy.number_format = copy(cell_above.number_format)
            cell_to_copy.protection = copy(cell_above.protection)
            cell_to_copy.alignment = copy(cell_above.alignment)


def table_testing():
    wb = openpyxl.load_workbook('tables/Mzdové náklady 2023-open.xlsx')
    ws = wb.worksheets[0]
    insert_row(ws, 20)
    wb.save('temp.xlsx')


def update_formulas():
    wb = openpyxl.load_workbook('tables/Mzdové náklady 2023-open.xlsx')
    ws = wb.worksheets[8]
    for j in range(3, 11):
        table_num = 1
        if j == 4:
            table_num = 3
        elif j == 5:
            table_num = 4
        elif j == 6:
            table_num = 7
        elif j == 7:
            table_num = 8
        elif j == 8:
            table_num = 2
        elif j == 10:
            table_num = 11
        for i in range(1, 13):
            if j == 9:
                string = f'=SUMIF(Tabulka9[přičítat{i}],"ANO",Tabulka9[součet{i}])'
            else:
                if i == 1:
                    string = f'=SUM(Tabulka{table_num}[součet])'
                else:
                    string = f'=SUM(Tabulka{table_num}[součet{i}])'
            ws.cell(j, i+1).value = string
    wb.save('tables/Mzdové náklady 2023-open.xlsx')


def update_f(column_name, tables, row):
    wb = openpyxl.load_workbook('tables/Mzdové náklady 2023-open.xlsx')
    ws = wb.worksheets[8]
    for col in range(1, 13):
        delimiter = ','
        if col == 1:
            num = ''
        else:
            num = col
        formula = f'=SUM('
        for table_num in tables:
            if table_num == tables[-1]:
                delimiter = ''
            formula += f'SUM(Tabulka{table_num}[{column_name}{num}]){delimiter}'
        formula += f')'
        print(formula)
        ws.cell(row, col + 1).value = formula

    wb.save('tables/Mzdové náklady 2023-open.xlsx')


def is_tool(name):
    return which(name) is not None


def saveas_libreoffice(file):

    if is_tool('libreoffice'):
        xlsx_tool = 'libreoffice'
    else:
        print('LibreOffice is not installed.')
        return 0

    # check if GTK_PATH is set, if yes, unset it
    if 'GTK_PATH' in os.environ:
        del os.environ['GTK_PATH']

    # Open xlsx editing software
    cmd = [xlsx_tool, file]
    subprocess.Popen(cmd)

    # Wait until spreadsheet is loaded
    time.sleep(3)

    # Switch focus to edit software
    cmd = ['wmctrl', '-a', xlsx_tool]
    subprocess.Popen(cmd)

    # Save as
    pyautogui.hotkey('ctrl', 'shift', 's')
    time.sleep(0.5)
    pyautogui.press('enter')
    time.sleep(0.5)
    pyautogui.press('right')
    time.sleep(0.5)
    pyautogui.press('enter')
    time.sleep(0.5)

    # Close the program
    pyautogui.hotkey('alt', 'f4')
    time.sleep(0.5)


def saveas_excel(filename):

    try:
        import win32com.client as win32 # type: ignore
        import pythoncom
    except ImportError:
        print('win32com is not installed. Please install it to use this script.')
        win32 = None
        exit()

    try:
        excel = win32.Dispatch('Excel.Application', pythoncom.CoInitialize())
        excel.Visible = False

        wb = excel.Workbooks.Open(Path(filename))
        wb.Close(True)

        excel.Application.Quit()
    except Exception as e:
        print(e)
        print('There is no excel installed.')

# def save_uploaded_file(f):
#     """ Saves uploaded file to temp directory """
#     with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as destination:
#         for chunk in f.chunks():
#             destination.write(chunk)
#         return destination.name

def save_uploaded_file(f, location):
    """Saves uploaded file to a specified location with a specified file name"""
    dirpath = Path(location)

    if not dirpath.exists():
        dirpath.mkdir()

    file_path = dirpath / str(f)

    with file_path.open('wb') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    return str(file_path)

# update_f('odměna', [1, 2, 3, 4, 7, 8, 9, 11], 11)
# update_f('str.', [1, 2, 3], 12)
# update_f('refundace', [1, 2, 3, 4, 7, 8], 13)
# update_f('jízdné', [1, 2, 3, 4, 7, 8], 14)
# update_f('mzd.nákl.', [1, 2, 3, 4, 7, 8, 9, 11], 15)
