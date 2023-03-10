from itertools import chain

import servant
from vanguard import Vanguard
from servant import mapping, column_map, split_name
from openpyxl.utils import get_column_letter


def to_list(idnum, items, row_num, num):
    message = f'{idnum} {items["Name"]} is in sheet{num} row #{row_num}.'
    for month, payout_data in items["Date"].items():
        m = (int(month.split('.')[0]) + 2) % 3

        month_col = mapping[num-2][m]

        message += f'\n-Payout for month {month}: {payout_data["Payout"]}=>{month_col}{row_num}'

    print(message)


def write_to_list(idnum, items, row_num, worksheet, new=False):
    # print(items)
    if worksheet.title.startswith('2'):
        num = 0
        offset = 1
    else:
        num = 1
        offset = 0

    for month, payout_data in items["Date"].items():
        m = (int(month.split('.')[0]) + 2) % 3
        shift = 14 - m
        if items["PensionType"]:
            month_col_status = column_map[num][m]
            worksheet.cell(row_num, month_col_status).value = items["PensionType"]

        month_col_payout = column_map[num][m] + offset
        worksheet.cell(row_num, month_col_payout).value = payout_data["Payout"]
        if payout_data["Payout"] > 0:
            worksheet.cell(row_num, month_col_payout + shift).value = f'=14200-' \
                                                                      f'{get_column_letter(month_col_payout + 1)}' \
                                                                      f'{row_num}'

        if num == 0:
            worksheet.cell(row_num, 6).value = items["EndEmployment"]

        if new:
            lastname, firstname = split_name(items["Name"])
            worksheet.cell(row_num, 2).value = lastname
            worksheet.cell(row_num, 3).value = firstname
            if num == 0:
                worksheet.cell(row_num, 4).value = idnum
                worksheet.cell(row_num, 5).value = items["StartEmployment"]
                worksheet.cell(row_num, 7).value = items["InsCode"]
                worksheet.cell(row_num, 8).value = items["PensionStart"]
            elif num == 1:
                worksheet.cell(row_num, 4).value = idnum[:6] + '/' + idnum[6:]
                worksheet.cell(row_num, 5).value = '-\'\'-'
                worksheet.cell(row_num, 6).value = 'PA'
                worksheet.cell(row_num, 7).value = '100%'


def new_person(idnum, items):
    message = f'{idnum} {items["Name"]} is new and will be placed to list'
    pen_t = items["PensionType"]
    if pen_t:
        message += ' 2'
    else:
        message += ' 3'


class Enforcer:
    def __init__(self, dataset):
        self.data_up = dataset[0][0]
        self.data_lo = dataset[0][1]
        self.merged_up = dataset[1][0]
        self.merged_lo = dataset[1][1]
        self.x = dataset[-1]
        self.last_row = [row[1] + 1 for row in self.x.range]

    def display_data(self):
        for person, data in self.merged_up.items():
            if person in self.data_up[0]:
                to_list(person, data, self.data_up[0][person], 2)
            elif person in self.data_up[1]:
                to_list(person, data, self.data_up[1][person], 3)
            else:
                print(f'{person} {data["Name"]} is new')
                if data['PensionType'] != '':
                    to_list(person, data, self.last_row[0], 2)
                    self.last_row[0] += 1
                    # message += 'belongs to list2'
                else:
                    to_list(person, data, self.last_row[1], 3)
                    self.last_row[1] += 1
                    # message += 'belongs to list3'
                # print(message)

    def display_lo(self):
        for person, data in self.merged_lo.items():
            for i, sheet_data in enumerate(self.data_lo):
                if i < 3:
                    fare_shift = 4
                else:
                    fare_shift = 2

                if person in sheet_data:
                    # print(i, data, sheet_data[person])
                    message = f'Sheet: {i}, Line: {sheet_data[person]}, Person: {person}, {data["Code"]}, {data["Cat"]}'
                    for date, money in data['Date'].items():
                        col = self.x.get_month(date, i)
                        fare_col = col + fare_shift
                        col_letter = get_column_letter(col)
                        fare_letter = get_column_letter(fare_col)
                        message += f'\n- {col_letter}{sheet_data[person]}: {money["Payout"]}'
                        if money["Fare"]:
                            message += f', {fare_letter}{sheet_data[person]}: {money["Fare"]}'
                    print(message)

    def write_data_lo(self):
        wb = self.x.wb_lo
        self.write_new_emps(wb)
        for person, data in self.merged_lo.items():
            for i, sheet_data in enumerate(self.data_lo):
                if i < 3:
                    fare_shift = 4
                else:
                    fare_shift = 2

                if person in sheet_data:
                    # print(i, data, sheet_data[person])
                    message = f'Sheet: {i}, Line: {sheet_data[person]}, Person: {person}, {data["Code"]}, {data["Cat"]}'
                    ws = wb.worksheets[i]
                    for date, money in data['Date'].items():
                        col = self.x.get_month(date, i)
                        fare_col = col + fare_shift
                        col_letter = get_column_letter(col)
                        fare_letter = get_column_letter(fare_col)
                        message += f'\n- {col_letter}{sheet_data[person]}: {money["Payout"]}'
                        ws.cell(sheet_data[person], col).value = money["Payout"]
                        if money["Fare"]:
                            message += f', {fare_letter}{sheet_data[person]}: {money["Fare"]}'
                            ws.cell(sheet_data[person], fare_col).value = money["Fare"]
        wb.save('temp.xlsx')

    def get_new_emps(self):
        # merged_lo is dictionary with all data for each employee
        # 'Arvensisov?? Radka': {'Name': 'Arvensisov?? Radka', 'Code': 'Prode',...
        # data_lo is list of dictionaries with key:value pairs being 'name':'line_number' for each sheet
        # [{'Bobok Vil??m': 3, 'Cenefels Jan': 4, 'Dbal?? Petr': 5, 'Diviak Miroslav': 6, ...
        data_for_month = set(self.merged_lo.keys())
        names_in_xlsx = set(chain.from_iterable(d.keys() for d in self.data_lo))
        # print('People without pay for this month')
        # print(names_in_xlsx.difference(data_for_month))
        # print(self.get_last_rows())
        # print('New people')
        return [self.merged_lo[name] for name in data_for_month.difference(names_in_xlsx)]

    def display_new_emps(self):
        print('New people bellow')
        last_lines = self.get_last_rows()
        for emp in enforcer.get_new_emps():
            sheet_index = servant.get_sheet_by_emp_data(emp['Code'], emp['Cat'])
            last_lines[sheet_index] += 1
            if sheet_index < 3:
                fare_shift = 4
            else:
                fare_shift = 2

            message = f'Sheet: {sheet_index}, Line: {last_lines[sheet_index]}, Person: {emp["Name"]}, ' \
                      f'{emp["Code"]}, {emp["Cat"]}'
            for date, money in emp['Date'].items():
                col = self.x.get_month(date, sheet_index)
                fare_col = col + fare_shift
                col_letter = get_column_letter(col)
                fare_letter = get_column_letter(fare_col)
                message += f'\n- {col_letter}{last_lines[sheet_index]}: {money["Payout"]}'
                if money["Fare"]:
                    message += f', {fare_letter}{last_lines[sheet_index]}: {money["Fare"]}'
            print(message)

    def write_new_emps(self, wb):
        last_lines = self.get_last_rows()

        for emp in enforcer.get_new_emps():
            sheet_index = servant.get_sheet_by_emp_data(emp['Code'], emp['Cat'])

            ws = wb.worksheets[sheet_index]
            servant.insert_row(ws, last_lines[sheet_index] + 1)
            ws.cell(last_lines[sheet_index] + 1, 2).value = emp["Name"]

            if emp["PensionType"]:
                ws.cell(last_lines[sheet_index] + 1, 1).value = 1
            else:
                ws.cell(last_lines[sheet_index] + 1, 1).value = 0

            if sheet_index < 3:
                fare_shift = 4
            else:
                fare_shift = 2

            for date, money in emp['Date'].items():
                col = self.x.get_month(date, sheet_index)
                fare_col = col + fare_shift
                ws.cell(last_lines[sheet_index] + 1, col).value = money["Payout"]
                if money["Fare"]:
                    ws.cell(last_lines[sheet_index] + 1, fare_col).value = money["Fare"]
            last_lines[sheet_index] += 1
        # wb.save('temp.xlsx')

    def write_data(self):
        wb = self.x.wb_up
        for person, data in self.merged_up.items():
            if person in self.data_up[0]:
                ws = wb.worksheets[1]
                write_to_list(person, data, self.data_up[0][person], ws)

            elif person in self.data_up[1]:
                ws = wb.worksheets[2]
                write_to_list(person, data, self.data_up[1][person], ws)
            else:
                # print(f'{person} {data["Name"]} is new')
                if data['PensionType'] != '':
                    ws = wb.worksheets[1]
                    # to_list(person, data, self.last_row[0], 2)
                    write_to_list(person, data, self.last_row[0], ws, new=True)
                    self.last_row[0] += 1
                    # message += 'belongs to list2'
                else:
                    # to_list(person, data, self.last_row[1], 3)
                    ws = wb.worksheets[2]
                    write_to_list(person, data, self.last_row[1], ws, new=True)
                    self.last_row[1] += 1
                    # message += 'belongs to list3'
                # print(message)
        wb.save('temp-up.xlsx')

    def get_last_rows(self):
        last_rows = [self.x.last_row_lo(i) for i in range(8)]
        return last_rows


vanguard = Vanguard(file_mzdy='data/Q2.CSV', file_pracov='data/PRACOVQ2.CSV')
enforcer = Enforcer(vanguard.loader)
# enforcer.display_data()
# enforcer.display_lo()
# enforcer.display_new_emps()
enforcer.write_data()
enforcer.write_data_lo()
