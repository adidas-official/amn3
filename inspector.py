import servant

import openpyxl
from openpyxl.utils import column_index_from_string

from pathlib import Path
import pandas as pd


class Inspector:
    def __init__(self, up_spreadsheet, lo_spreadsheet):
        self.file_up = up_spreadsheet
        self.file_lo = lo_spreadsheet
        self.wb_up, self.wb_lo = self.is_readable()
        self.q = self.quarter

    def is_readable(self):
        wb_up = openpyxl.load_workbook(self.file_up, data_only=True)
        ws = wb_up.worksheets[0]

        if ws.cell(21, 7).value is None:
            servant.resave_xlsx(self.file_up)
            wb_up = openpyxl.load_workbook(self.file_up, data_only=True)

        wb_lo = servant.unlock(self.file_lo, '13881744', data_only=True)
        ws = wb_lo.worksheets[-2]

        if ws.cell(3, 2).value is None:
            servant.resave_xlsx(self.file_lo)
            wb_lo = servant.unlock(self.file_lo, '13881744', data_only=True)

        return wb_up, wb_lo

    @property
    def refund_up(self):

        ws = self.wb_up.worksheets[0]
        ref_val = ws.cell(21, 7).value

        return ref_val

    @property
    def refund_lo(self):
        ws = self.wb_lo.worksheets[-2]

        if self.q:
            ref_val = ws.cell(self.q + 2, column_index_from_string('R')).value
            return ref_val
        else:
            print(f'Quarter is not defined. Check file {self.file_up}, cell D4')

    @property
    def quarter(self):
        ws = self.wb_up.worksheets[0]
        quarter = ws.cell(6, 4).value

        try:
            return int(quarter)
        except ValueError:
            print(f'Quarter is not integer type. Fix the value in {self.file_up} in cell D6')
            return False

    def refund_are_equal(self):
        rup = self.refund_up
        rlo = self.refund_lo
        if rup == rlo:
            return True
        else:
            print('Refunds are not equal')
            print(f'RUP: {rup}, RLO: {rlo}, difference: {rup - rlo}')
            return False

    def get_refund_of_emp(self, row, month):
        ws = self.wb_up.worksheets[1]
        total = 0

        for i in range(3):
            val = ws.cell(row, 12 + (month * 5) + i).value
            if val:
                total += val
        return total

    def sum_month_refund(self, month):
        ws = self.wb_up.worksheets[1]

        row = 13
        total = 0

        while True:
            if not ws.cell(row, 2).value:
                break

            total += self.get_refund_of_emp(row, month)
            row += 1
        return total

    @property
    def sums_of_refunds(self):
        months = [self.sum_month_refund(i) for i in range(3)]
        return months

    def refund_for_month_lo(self, month):
        month_index = self.q * 3 - 2 + month
        ws = self.wb_lo.worksheets[-2]
        return ws.cell(13, month_index + 1).value

    @property
    def sums_of_refunds_lo(self):
        return [self.refund_for_month_lo(month) for month in range(3)]

    @property
    def faulty_months(self):
        if not self.refund_are_equal():
            faulty_months = []
            refunds_up = self.sums_of_refunds
            refunds_lo = self.sums_of_refunds_lo

            for month_index, month in enumerate(zip(refunds_up, refunds_lo)):
                if month[0] != month[1]:
                    faulty_months.append(month_index)

            return faulty_months
        return False

    def get_month(self, month, sheet_num):
        """ Gets month column index in local table from date in employee data object """
        # Counter for keeping track of curent column index
        counter = 0
        # Number of how many times merged cell was found
        m = 0
        month_num = self.q * 3 - 2 + month
        cell = False

        while True:
            if m == month_num:
                return cell.column

            cell = self.wb_lo.worksheets[sheet_num].cell(row=1, column=counter + 3)
            counter += 1

            if not type(cell).__name__ == 'MergedCell' and cell.value:
                m += 1

    @property
    def names_in_table(self):
        ws = self.wb_up.worksheets[1]
        names = {}

        row = 13

        while True:
            name = ws.cell(row, 2).value
            if not name:
                break

            names.setdefault(name, [])

            for i in range(3):
                total = self.get_refund_of_emp(row, i)
                names[name].append(total)
            row += 1
        return names

    def names_in_table_2(self):
        wb = openpyxl.load_workbook('temp.xlsx', data_only=True, read_only=False)
        sheets_df = []
        table_indexes = [1, 2, 3, 4, 7, 8]

        # Iterate over sheets
        for sheetindex, ws in zip(table_indexes, wb.worksheets[:-4]):
            headers = []
            tab = ws.tables[f'Tabulka{sheetindex}']
            table_range = tab.ref
            data_rows = []

            header = False

            for row in ws[table_range]:
                data_cols = []

                for cell in row:
                    if not header:
                        headers.append(cell.value)
                    else:
                        data_cols.append(cell.value)

                if not header:
                    header = True
                    continue

                data_rows.append(data_cols)

            df = pd.DataFrame(data_rows, columns=headers, index=None)
            print(df[['Jméno', 'refundace7']])


file_up = 'temp-up.xlsx'
file_lo = 'temp.xlsx'


if all((Path(file_up).exists(), Path(file_lo).exists())):
    inspector = Inspector(file_up, file_lo)
    inspector.names_in_table_2()
else:
    print('File does not exist')
